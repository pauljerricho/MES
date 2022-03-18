from cgi import print_arguments
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "SungilSheet"

#A1셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을 땐 'NONE'을 출력

# row = 1,2,3
# column = A,B,C,D
print(ws.cell(row=1, column=1).value) #A1값이 출력
print(ws.cell(row=1, column=2).value) #B1값이 출력



wb.save("sample2.xlsx")

