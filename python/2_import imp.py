from openpyxl import Workbook
wb = Workbook() #새 워크북 생성
ws = wb.active #현재 활성화 된 sheet를 가져옴
ws.title = "SungilSheet" #sheet의 이름을 바꿈
ws.sheet_properties.tabColor = "ff56ff" #RGB 형태의 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NEWSHEET" , 2) # 차트 생성 2번째 인덱스에  시트 생성
new_ws = wb["NEWSHEET"] # DICT 형태로 SHEET에 접근

print(wb.sheetnames)

# Sheet 복사
new_ws["A1"] = "TEST"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
#RPA PROJECT(사무자동화)