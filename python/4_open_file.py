from openpyxl import load_workbook
wb = load_workbook("sample2.xlsx") #sample.xlsx 파일에서 wb 불러옴
ws = wb.active

for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column=y).value, end =" ") # 1 2 3 4..
    
    print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x,column=y).value, end =" ") # 1 2 3 4..
    
    print()
#RPA PROJECT(사무자동화)